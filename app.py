from flask import Flask, render_template, request, jsonify, redirect, url_for, session
import json
import os
import openpyxl
import pandas as pd
from functools import wraps
from datetime import datetime
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'clave_por_defecto_para_desarrollo')

# ============ CONFIGURACIÓN DE CARGA DE ARCHIVOS ============

UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv'}

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs('data', exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 10 * 1024 * 1024  # 10MB

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# ============ DATOS DE USUARIOS ============

USERS = {
    'admin': {
        'password': 'admin123',
        'role': 'admin',
        'name': 'Administrador'
    },
    '28233037': {
        'password': 'agustin2026',
        'role': 'student',
        'name': 'AGUSTIN ANDRADE',
        'email': 'agusuniversidad219@gmail.com'
    },
    '27111264': {
        'password': 'mauricio2026',
        'role': 'student',
        'name': 'MAURICIO LEAL',
        'email': 'mleal4c@gmail.com'
    },
    '21073829': {
        'password': 'maria2026',
        'role': 'student',
        'name': 'MARIA DANIELA MEDINA',
        'email': 'ma.danielamedina@gmail.com'
    },
    '29529906': {
        'password': 'luis2026',
        'role': 'student',
        'name': 'LUIS RIOS',
        'email': 'lavidusriya@gmail.com'
    },
    '29501138': {
        'password': 'daniela2026',
        'role': 'student',
        'name': 'DANIELA RONDON',
        'email': 'rondonbrazondaniela@gmail.com'
    }
}

# ============ DECORADORES ============

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session or session.get('role') != 'admin':
            return jsonify({'error': 'Acceso denegado'}), 403
        return f(*args, **kwargs)
    return decorated_function

# ============ FUNCIONES DE CÁLCULO ============

def get_safe_value(cell):
    """Obtiene el valor de una celda manejando diferentes tipos"""
    valor = cell.value
    if valor is None:
        return 0
    if isinstance(valor, str):
        if valor.startswith('='):
            return 0
        try:
            return float(valor)
        except:
            return 0
    try:
        return float(valor)
    except:
        return 0

def calcular_promedio_semanas(notas, semanas_totales=10):
    """Calcula el promedio de todas las semanas que tienen notas (ignora ceros)"""
    if not notas:
        return 0
    notas_validas = [n for n in notas[:semanas_totales] if n > 0]
    if not notas_validas:
        return 0
    return sum(notas_validas) / len(notas_validas)

def contar_semanas_evaluadas(notas, semanas_totales=10):
    """Cuenta cuántas semanas tienen nota (valor > 0)"""
    if not notas:
        return 0
    return len([n for n in notas[:semanas_totales] if n > 0])

def calcular_nota_cuaderno(notas):
    return calcular_promedio_semanas(notas, 10)

def calcular_nota_gestion(notas):
    return calcular_promedio_semanas(notas, 10)

def calcular_nota_quiz(notas):
    return calcular_promedio_semanas(notas, 10)

def calcular_nota_informe(informe_notas):
    """Calcula la nota del informe (promedio de todos los informes disponibles)"""
    if not informe_notas:
        return 0
    # Filtrar solo informes con nota > 0
    notas_validas = [n for n in informe_notas if n > 0]
    if not notas_validas:
        return 0
    return sum(notas_validas) / len(notas_validas)

def calcular_nota_bloqueA(cuaderno, gestion, quiz, informe, ponderaciones):
    nota_cuaderno = calcular_nota_cuaderno(cuaderno)
    nota_gestion = calcular_nota_gestion(gestion)
    nota_quiz = calcular_nota_quiz(quiz)
    nota_informe = calcular_nota_informe(informe)
    
    nota_ponderada = (
        nota_cuaderno * ponderaciones.get('cuaderno_pct', 0.20) +
        nota_gestion * ponderaciones.get('gestion_pct', 0.15) +
        nota_quiz * ponderaciones.get('quiz_pct', 0.25) +
        nota_informe * ponderaciones.get('informe_pct', 0.40)
    )
    return round(nota_ponderada, 2)

def calcular_nota_final(bloqueA, bloqueB, ponderacion_bloqueA=0.85, ponderacion_bloqueB=0.15):
    nota_final = (bloqueA * ponderacion_bloqueA) + (bloqueB * ponderacion_bloqueB)
    return round(nota_final, 2)

# ============ FUNCIONES PARA PROCESAR LISTADO DE ESTUDIANTES ============

def procesar_listado_estudiantes(filepath):
    """Procesa el archivo Excel con el listado de estudiantes"""
    try:
        df = pd.read_excel(filepath, header=None)
        
        start_row = None
        for idx, row in df.iterrows():
            if row.astype(str).str.contains('CÉDULA', case=False, na=False).any():
                start_row = idx + 1
                break
        
        if start_row is None:
            for idx, row in df.iterrows():
                if isinstance(row[1], str) and 'QUIMICA' in row[1]:
                    start_row = idx
                    break
        
        if start_row is None:
            return {'error': 'No se pudo identificar el formato del archivo'}
        
        estudiantes = []
        for idx in range(start_row, len(df)):
            row = df.iloc[idx]
            cedula = row[2] if len(row) > 2 else None
            if not cedula or str(cedula).strip() == '' or str(cedula) == 'nan':
                continue
            
            try:
                cedula_str = str(int(cedula)).strip() if isinstance(cedula, (int, float)) else str(cedula).strip()
                nombres = str(row[3]).strip() if len(row) > 3 and not pd.isna(row[3]) else ''
                apellidos = str(row[4]).strip() if len(row) > 4 and not pd.isna(row[4]) else ''
                correo = str(row[5]).strip() if len(row) > 5 and not pd.isna(row[5]) else ''
                
                nombre_completo = f"{nombres} {apellidos}".strip()
                
                estudiante = {
                    'cedula': cedula_str,
                    'nombres': nombres,
                    'apellidos': apellidos,
                    'nombre_completo': nombre_completo,
                    'correo': correo
                }
                estudiantes.append(estudiante)
            except Exception as e:
                print(f"Error al procesar fila {idx}: {e}")
                continue
        
        return {
            'success': True,
            'estudiantes': estudiantes,
            'total': len(estudiantes)
        }
    except Exception as e:
        return {'error': f'Error al procesar archivo: {str(e)}'}

def actualizar_usuarios_desde_listado(estudiantes, password_base='estudiante2026'):
    """Actualiza el diccionario USERS con los estudiantes del listado"""
    nuevos_usuarios = {}
    for estudiante in estudiantes:
        cedula = estudiante['cedula']
        nombre_completo = estudiante['nombre_completo']
        correo = estudiante['correo']
        nuevos_usuarios[cedula] = {
            'password': password_base,
            'role': 'student',
            'name': nombre_completo,
            'email': correo,
            'cedula': cedula
        }
    return nuevos_usuarios

# ============ CARGA DE DATOS DESDE EXCEL ============

def cargar_datos_excel():
    try:
        wb = openpyxl.load_workbook('Calificaciones I-2026.xlsx', data_only=True)
        datos = {}
        
        estudiantes = list(USERS.keys())
        estudiantes = [USERS[e]['name'] for e in estudiantes if USERS[e]['role'] == 'student']
        
        # Cargar ponderaciones
        ws_pond = wb['Ponderaciones']
        ponderaciones = {
            'cuaderno_pct': get_safe_value(ws_pond.cell(row=7, column=3)) / 100,
            'gestion_pct': get_safe_value(ws_pond.cell(row=8, column=3)) / 100,
            'quiz_pct': get_safe_value(ws_pond.cell(row=9, column=3)) / 100,
            'informe_pct': get_safe_value(ws_pond.cell(row=10, column=3)) / 100,
            'bloqueA_pct': get_safe_value(ws_pond.cell(row=3, column=3)) / 100,
            'bloqueB_pct': get_safe_value(ws_pond.cell(row=4, column=3)) / 100
        }
        datos['ponderaciones'] = ponderaciones
        
        # Inicializar estructuras
        datos['cuaderno'] = {}
        datos['gestion'] = {}
        datos['quiz'] = {}
        datos['informes'] = {}
        datos['bloqueA'] = {}
        datos['bloqueB_raw'] = {}
        datos['final'] = {}
        datos['semanas_evaluadas'] = {}
        
        # Cargar datos por estudiante
        for estudiante in estudiantes:
            # Buscar fila del estudiante en cada hoja
            # Cuaderno
            ws_cuaderno = wb['Cuaderno']
            for i in range(16, 25):
                if ws_cuaderno.cell(row=i, column=2).value == estudiante:
                    notas = []
                    for j in range(3, 13):
                        notas.append(get_safe_value(ws_cuaderno.cell(row=i, column=j)))
                    datos['cuaderno'][estudiante] = notas
                    break
            
            # Gestión
            ws_gestion = wb['Gestión del laboratorio']
            for i in range(4, 12):
                if ws_gestion.cell(row=i, column=1).value == estudiante:
                    notas = []
                    for j in range(3, 13):
                        notas.append(get_safe_value(ws_gestion.cell(row=i, column=j)))
                    datos['gestion'][estudiante] = notas
                    break
            
            # Quiz
            ws_quiz = wb['Quiz']
            for i in range(5, 12):
                if ws_quiz.cell(row=i, column=1).value == estudiante:
                    notas = []
                    for j in range(3, 13):
                        notas.append(get_safe_value(ws_quiz.cell(row=i, column=j)))
                    datos['quiz'][estudiante] = notas
                    break
            
            # Informes
            ws_informes = wb['Notas informes']
            for i in range(4, 12):
                if ws_informes.cell(row=i, column=1).value == estudiante:
                    notas = []
                    for j in range(3, 8):
                        notas.append(get_safe_value(ws_informes.cell(row=i, column=j)))
                    datos['informes'][estudiante] = notas
                    break
            
            # Calcular Bloque A
            cuaderno = datos['cuaderno'].get(estudiante, [0]*10)
            gestion = datos['gestion'].get(estudiante, [0]*10)
            quiz = datos['quiz'].get(estudiante, [0]*10)
            informe = datos['informes'].get(estudiante, [0]*5)
            
            datos['semanas_evaluadas'][estudiante] = {
                'cuaderno': contar_semanas_evaluadas(cuaderno, 10),
                'gestion': contar_semanas_evaluadas(gestion, 10),
                'quiz': contar_semanas_evaluadas(quiz, 10)
            }
            
            bloqueA = calcular_nota_bloqueA(
                cuaderno, gestion, quiz, informe,
                {
                    'cuaderno_pct': ponderaciones['cuaderno_pct'],
                    'gestion_pct': ponderaciones['gestion_pct'],
                    'quiz_pct': ponderaciones['quiz_pct'],
                    'informe_pct': ponderaciones['informe_pct']
                }
            )
            datos['bloqueA'][estudiante] = bloqueA
            
            # Bloque B (Seminario)
            datos['bloqueB_raw'][estudiante] = 0
            
            # Nota Final
            datos['final'][estudiante] = calcular_nota_final(
                bloqueA, 0,
                ponderaciones['bloqueA_pct'],
                ponderaciones['bloqueB_pct']
            )
        
        print("\n--- DATOS CARGADOS (Base 20 - 10 semanas) ---")
        for estudiante in estudiantes:
            print(f"{estudiante}: Bloque A: {datos['bloqueA'].get(estudiante, 0):.2f}, Final: {datos['final'].get(estudiante, 0):.2f}")
        
        return datos
    
    except FileNotFoundError:
        print("El archivo Excel no se encuentra. Usando datos de ejemplo.")
        return generar_datos_ejemplo()
    except Exception as e:
        print(f"Error al cargar Excel: {e}")
        import traceback
        traceback.print_exc()
        return generar_datos_ejemplo()

def generar_datos_ejemplo():
    estudiantes = [USERS[e]['name'] for e in USERS if USERS[e]['role'] == 'student']
    datos = {
        'cuaderno': {},
        'gestion': {},
        'quiz': {},
        'informes': {},
        'bloqueA': {},
        'bloqueB_raw': {},
        'final': {},
        'semanas_evaluadas': {},
        'ponderaciones': {
            'cuaderno_pct': 0.20,
            'gestion_pct': 0.15,
            'quiz_pct': 0.25,
            'informe_pct': 0.40,
            'bloqueA_pct': 0.85,
            'bloqueB_pct': 0.15
        }
    }
    
    for estudiante in estudiantes:
        datos['cuaderno'][estudiante] = [15.5, 18.0, 19.0, 17.0, 0, 0, 0, 0, 0, 0]
        datos['gestion'][estudiante] = [18.0, 19.0, 17.0, 18.0, 0, 0, 0, 0, 0, 0]
        datos['quiz'][estudiante] = [17.0, 18.0, 16.0, 17.5, 0, 0, 0, 0, 0, 0]
        datos['informes'][estudiante] = [15.0, 0, 0, 0, 0]
        datos['semanas_evaluadas'][estudiante] = {'cuaderno': 4, 'gestion': 4, 'quiz': 4}
        
        bloqueA = calcular_nota_bloqueA(
            datos['cuaderno'][estudiante],
            datos['gestion'][estudiante],
            datos['quiz'][estudiante],
            datos['informes'][estudiante],
            datos['ponderaciones']
        )
        datos['bloqueA'][estudiante] = bloqueA
        datos['bloqueB_raw'][estudiante] = 0
        datos['final'][estudiante] = bloqueA * 0.85
    
    return datos

# ============ FILTROS PARA FORMATO DECIMAL CON COMA ============

@app.template_filter('comma')
def format_with_comma(value):
    if value is None:
        return "0,00"
    try:
        num = float(value)
        return f"{num:.2f}".replace('.', ',')
    except (ValueError, TypeError):
        return "0,00"

@app.template_filter('comma_auto')
def format_with_comma_auto(value):
    if value is None:
        return "0"
    try:
        num = float(value)
        if num == int(num):
            return f"{int(num)}"
        return f"{num:.2f}".replace('.', ',')
    except (ValueError, TypeError):
        return "0"

app.jinja_env.filters['comma'] = format_with_comma
app.jinja_env.filters['comma_auto'] = format_with_comma_auto

# ============ CARGA INICIAL DE DATOS ============

DATOS = cargar_datos_excel()

def guardar_datos_json():
    with open('data/calificaciones.json', 'w', encoding='utf-8') as f:
        json.dump(DATOS, f, ensure_ascii=False, indent=2)

guardar_datos_json()

# ============ INICIALIZAR ESTRUCTURAS PARA TODOS LOS ESTUDIANTES ============

def inicializar_estructuras_para_todos():
    """Asegura que cada estudiante tenga datos en todas las secciones"""
    estudiantes = [USERS[e]['name'] for e in USERS if USERS[e]['role'] == 'student']
    
    for estudiante in estudiantes:
        # Cuaderno (10 semanas)
        if estudiante not in DATOS['cuaderno']:
            DATOS['cuaderno'][estudiante] = [0.0] * 10
            
        # Gestión (10 semanas)
        if estudiante not in DATOS['gestion']:
            DATOS['gestion'][estudiante] = [0.0] * 10
            
        # Quiz (10 semanas)
        if estudiante not in DATOS['quiz']:
            DATOS['quiz'][estudiante] = [0.0] * 10
            
        # Informes (5 informes A-E)
        if estudiante not in DATOS['informes']:
            DATOS['informes'][estudiante] = [0.0] * 5
            
        # Bloque A
        if estudiante not in DATOS['bloqueA']:
            DATOS['bloqueA'][estudiante] = 0.0
            
        # Bloque B
        if estudiante not in DATOS['bloqueB_raw']:
            DATOS['bloqueB_raw'][estudiante] = 0.0
            
        # Nota Final
        if estudiante not in DATOS['final']:
            DATOS['final'][estudiante] = 0.0
            
        # Semanas evaluadas
        if estudiante not in DATOS['semanas_evaluadas']:
            DATOS['semanas_evaluadas'][estudiante] = {
                'cuaderno': 0,
                'gestion': 0,
                'quiz': 0
            }
    
    print(f"✅ Estructuras inicializadas para {len(estudiantes)} estudiantes")
    print(f"📊 Cuaderno: {len(DATOS['cuaderno'])} estudiantes")
    print(f"📊 Gestión: {len(DATOS['gestion'])} estudiantes")
    print(f"📊 Quiz: {len(DATOS['quiz'])} estudiantes")
    print(f"📊 Informes: {len(DATOS['informes'])} estudiantes")

# Ejecutar la inicialización
inicializar_estructuras_para_todos()

# Guardar los cambios
guardar_datos_json()

# ============ RUTAS PRINCIPALES ============

@app.route('/')
def index():
    if 'user_id' in session:
        if session.get('role') == 'admin':
            return redirect(url_for('admin'))
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        if username in USERS and USERS[username]['password'] == password:
            session['user_id'] = username
            session['user_name'] = USERS[username]['name']
            session['role'] = USERS[username]['role']
            
            if session['role'] == 'admin':
                return redirect(url_for('admin'))
            return redirect(url_for('dashboard'))
        
        return render_template('login.html', error='Usuario o contraseña incorrectos')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ============ RUTAS DE ESTUDIANTE ============

@app.route('/dashboard')
@login_required
def dashboard():
    if session.get('role') == 'admin':
        return redirect(url_for('admin'))
    
    estudiante = session.get('user_name')
    cedula = session.get('user_id')
    
    datos_estudiante = {
        'nombre': estudiante,
        'cedula': cedula,
        'cuaderno': DATOS['cuaderno'].get(estudiante, [0]*10),
        'gestion': DATOS['gestion'].get(estudiante, [0]*10),
        'quiz': DATOS['quiz'].get(estudiante, [0]*10),
        'informes': DATOS['informes'].get(estudiante, [0]*5),
        'bloqueA': DATOS['bloqueA'].get(estudiante, 0),
        'bloqueB': DATOS['bloqueB_raw'].get(estudiante, 0),
        'final': DATOS['final'].get(estudiante, 0),
        'semanas': DATOS['semanas_evaluadas'].get(estudiante, {'cuaderno': 0, 'gestion': 0, 'quiz': 0})
    }
    
    datos_estudiante['cuaderno_prom'] = calcular_nota_cuaderno(datos_estudiante['cuaderno'])
    datos_estudiante['gestion_prom'] = calcular_nota_gestion(datos_estudiante['gestion'])
    datos_estudiante['quiz_prom'] = calcular_nota_quiz(datos_estudiante['quiz'])
    datos_estudiante['informe_sum'] = calcular_nota_informe(datos_estudiante['informes'])
    
    return render_template('dashboard.html', estudiante=datos_estudiante)

# ============ RUTAS DE ADMINISTRADOR ============

@app.route('/admin')
@login_required
@admin_required
def admin():
    return render_template('admin.html', estudiantes=DATOS, enumerate=enumerate)

@app.route('/admin/cargar-listado', methods=['GET', 'POST'])
@login_required
@admin_required
def cargar_listado():
    if request.method == 'POST':
        if 'file' not in request.files:
            return jsonify({'error': 'No se seleccionó ningún archivo'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'Nombre de archivo vacío'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'error': 'Tipo de archivo no permitido. Use .xlsx, .xls o .csv'}), 400
        
        try:
            filename = secure_filename(f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}")
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            
            resultado = procesar_listado_estudiantes(filepath)
            
            if not resultado.get('success'):
                return jsonify({'error': resultado.get('error', 'Error al procesar')}), 400
            
            estudiantes = resultado['estudiantes']
            nuevos_usuarios = actualizar_usuarios_desde_listado(estudiantes)
            
            global USERS
            for cedula, datos in nuevos_usuarios.items():
                USERS[cedula] = datos
            
            with open('data/usuarios_backup.json', 'w', encoding='utf-8') as f:
                json.dump(USERS, f, ensure_ascii=False, indent=2)
            
            for estudiante in estudiantes:
                nombre_completo = estudiante['nombre_completo']
                if nombre_completo not in DATOS['cuaderno']:
                    DATOS['cuaderno'][nombre_completo] = [0] * 10
                if nombre_completo not in DATOS['gestion']:
                    DATOS['gestion'][nombre_completo] = [0] * 10
                if nombre_completo not in DATOS['quiz']:
                    DATOS['quiz'][nombre_completo] = [0] * 10
                if nombre_completo not in DATOS['informes']:
                    DATOS['informes'][nombre_completo] = [0] * 5
                if nombre_completo not in DATOS['bloqueA']:
                    DATOS['bloqueA'][nombre_completo] = 0
                if nombre_completo not in DATOS['bloqueB_raw']:
                    DATOS['bloqueB_raw'][nombre_completo] = 0
                if nombre_completo not in DATOS['final']:
                    DATOS['final'][nombre_completo] = 0
                if nombre_completo not in DATOS['semanas_evaluadas']:
                    DATOS['semanas_evaluadas'][nombre_completo] = {
                        'cuaderno': 0,
                        'gestion': 0,
                        'quiz': 0
                    }
            
            guardar_datos_json()
            
            return jsonify({
                'success': True,
                'total_estudiantes': len(estudiantes),
                'estudiantes': estudiantes,
                'mensaje': f'✅ {len(estudiantes)} estudiantes cargados correctamente'
            })
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return jsonify({'error': str(e)}), 500
    
    return render_template('cargar_listado.html')

@app.route('/admin/preview-listado', methods=['POST'])
@login_required
@admin_required
def preview_listado():
    if 'file' not in request.files:
        return jsonify({'error': 'No se seleccionó ningún archivo'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Nombre de archivo vacío'}), 400
    
    if not allowed_file(file.filename):
        return jsonify({'error': 'Tipo de archivo no permitido'}), 400
    
    try:
        filename = secure_filename(f"preview_{file.filename}")
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        resultado = procesar_listado_estudiantes(filepath)
        
        if os.path.exists(filepath):
            os.remove(filepath)
        
        return jsonify(resultado)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/admin/estudiantes')
@login_required
@admin_required
def listar_estudiantes():
    estudiantes = []
    for cedula, datos in USERS.items():
        if datos['role'] == 'student':
            estudiantes.append({
                'cedula': cedula,
                'nombre': datos['name'],
                'correo': datos.get('email', ''),
                'nota_final': DATOS['final'].get(datos['name'], 0)
            })
    return jsonify(estudiantes)

@app.route('/admin/eliminar-estudiante/<cedula>', methods=['DELETE'])
@login_required
@admin_required
def eliminar_estudiante(cedula):
    if cedula not in USERS:
        return jsonify({'error': 'Estudiante no encontrado'}), 404
    
    nombre = USERS[cedula]['name']
    del USERS[cedula]
    
    for key in ['cuaderno', 'gestion', 'quiz', 'informes', 'bloqueA', 'bloqueB_raw', 'final', 'semanas_evaluadas']:
        if nombre in DATOS.get(key, {}):
            del DATOS[key][nombre]
    
    guardar_datos_json()
    return jsonify({'success': True, 'mensaje': f'Estudiante {nombre} eliminado'})

# ============ API ============

@app.route('/api/datos')
@login_required
def api_datos():
    if session.get('role') == 'admin':
        return jsonify(DATOS)
    
    estudiante = session.get('user_name')
    return jsonify({
        'nombre': estudiante,
        'datos': {
            'cuaderno': DATOS['cuaderno'].get(estudiante, []),
            'gestion': DATOS['gestion'].get(estudiante, []),
            'quiz': DATOS['quiz'].get(estudiante, []),
            'informes': DATOS['informes'].get(estudiante, []),
            'bloqueA': DATOS['bloqueA'].get(estudiante, 0),
            'bloqueB': DATOS['bloqueB_raw'].get(estudiante, 0),
            'final': DATOS['final'].get(estudiante, 0)
        }
    })

@app.route('/api/actualizar', methods=['POST'])
@login_required
@admin_required
def api_actualizar():
    data = request.json
    estudiante = data.get('estudiante')
    campo = data.get('campo')
    valor = data.get('valor')
    indice = data.get('indice', None)
    
    if not estudiante or not campo or valor is None:
        return jsonify({'error': 'Datos incompletos'}), 400
    
    try:
        valor = float(valor)
        if valor < 0 or valor > 20:
            return jsonify({'error': 'La nota debe estar entre 0 y 20'}), 400
        
        if campo == 'bloqueA':
            DATOS['bloqueA'][estudiante] = valor
            bloqueB = DATOS['bloqueB_raw'].get(estudiante, 0)
            DATOS['final'][estudiante] = calcular_nota_final(
                valor, bloqueB,
                DATOS['ponderaciones']['bloqueA_pct'],
                DATOS['ponderaciones']['bloqueB_pct']
            )
        elif campo == 'bloqueB':
            DATOS['bloqueB_raw'][estudiante] = valor
            bloqueA = DATOS['bloqueA'].get(estudiante, 0)
            DATOS['final'][estudiante] = calcular_nota_final(
                bloqueA, valor,
                DATOS['ponderaciones']['bloqueA_pct'],
                DATOS['ponderaciones']['bloqueB_pct']
            )
        elif campo in ['cuaderno', 'gestion', 'quiz']:
            if indice is not None:
                DATOS[campo][estudiante][indice] = valor
            else:
                DATOS[campo][estudiante] = [float(v) for v in valor]
            
            semanas = contar_semanas_evaluadas(DATOS[campo][estudiante], 10)
            if campo == 'cuaderno':
                DATOS['semanas_evaluadas'][estudiante]['cuaderno'] = semanas
            elif campo == 'gestion':
                DATOS['semanas_evaluadas'][estudiante]['gestion'] = semanas
            elif campo == 'quiz':
                DATOS['semanas_evaluadas'][estudiante]['quiz'] = semanas
            
            nuevo_bloqueA = calcular_nota_bloqueA(
                DATOS['cuaderno'][estudiante],
                DATOS['gestion'][estudiante],
                DATOS['quiz'][estudiante],
                DATOS['informes'][estudiante],
                DATOS['ponderaciones']
            )
            DATOS['bloqueA'][estudiante] = nuevo_bloqueA
            bloqueB = DATOS['bloqueB_raw'].get(estudiante, 0)
            DATOS['final'][estudiante] = calcular_nota_final(
                nuevo_bloqueA, bloqueB,
                DATOS['ponderaciones']['bloqueA_pct'],
                DATOS['ponderaciones']['bloqueB_pct']
            )
        elif campo == 'informes':
            if indice is not None:
                DATOS['informes'][estudiante][indice] = valor
            else:
                DATOS['informes'][estudiante] = [float(v) for v in valor]
            
            nuevo_bloqueA = calcular_nota_bloqueA(
                DATOS['cuaderno'][estudiante],
                DATOS['gestion'][estudiante],
                DATOS['quiz'][estudiante],
                DATOS['informes'][estudiante],
                DATOS['ponderaciones']
            )
            DATOS['bloqueA'][estudiante] = nuevo_bloqueA
            bloqueB = DATOS['bloqueB_raw'].get(estudiante, 0)
            DATOS['final'][estudiante] = calcular_nota_final(
                nuevo_bloqueA, bloqueB,
                DATOS['ponderaciones']['bloqueA_pct'],
                DATOS['ponderaciones']['bloqueB_pct']
            )
        else:
            return jsonify({'error': 'Campo no válido'}), 400
        
        guardar_datos_json()
        return jsonify({'success': True, 'nuevos_datos': {
            'bloqueA': DATOS['bloqueA'][estudiante],
            'final': DATOS['final'][estudiante]
        }})
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/estadisticas')
@login_required
@admin_required
def api_estadisticas():
    try:
        notas_finales = []
        estudiantes_lista = []
        
        for nombre, nota in DATOS['final'].items():
            nota_valor = float(nota) if nota is not None else 0
            notas_finales.append(nota_valor)
            estudiantes_lista.append({
                'nombre': nombre,
                'nota': round(nota_valor, 2),
                'bloqueA': round(float(DATOS['bloqueA'].get(nombre, 0)), 2),
                'bloqueB': round(float(DATOS['bloqueB_raw'].get(nombre, 0)), 2)
            })
        
        if not notas_finales:
            return jsonify({
                'promedio': 0,
                'maxima': 0,
                'minima': 0,
                'aprobados': 0,
                'reprobados': 0,
                'total': 0,
                'distribucion': [0, 0, 0, 0, 0],
                'estudiantes': []
            })
        
        promedio = round(sum(notas_finales) / len(notas_finales), 2)
        maxima = round(max(notas_finales), 2)
        minima = round(min(notas_finales), 2)
        
        aprobados = len([n for n in notas_finales if n >= 10])
        reprobados = len([n for n in notas_finales if n < 10])
        
        distribucion = [0, 0, 0, 0, 0]
        for nota in notas_finales:
            if nota < 5:
                distribucion[0] += 1
            elif nota < 10:
                distribucion[1] += 1
            elif nota < 15:
                distribucion[2] += 1
            elif nota < 18:
                distribucion[3] += 1
            else:
                distribucion[4] += 1
        
        return jsonify({
            'promedio': promedio,
            'maxima': maxima,
            'minima': minima,
            'aprobados': aprobados,
            'reprobados': reprobados,
            'total': len(notas_finales),
            'distribucion': distribucion,
            'estudiantes': estudiantes_lista
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/recalcular')
@login_required
@admin_required
def api_recalcular():
    try:
        ponderaciones = DATOS['ponderaciones']
        estudiantes = list(DATOS['cuaderno'].keys())
        
        for estudiante in estudiantes:
            cuaderno = DATOS['cuaderno'][estudiante]
            gestion = DATOS['gestion'][estudiante]
            quiz = DATOS['quiz'][estudiante]
            informe = DATOS['informes'][estudiante]
            
            DATOS['semanas_evaluadas'][estudiante] = {
                'cuaderno': contar_semanas_evaluadas(cuaderno, 10),
                'gestion': contar_semanas_evaluadas(gestion, 10),
                'quiz': contar_semanas_evaluadas(quiz, 10)
            }
            
            nuevo_bloqueA = calcular_nota_bloqueA(
                cuaderno, gestion, quiz, informe,
                {
                    'cuaderno_pct': ponderaciones['cuaderno_pct'],
                    'gestion_pct': ponderaciones['gestion_pct'],
                    'quiz_pct': ponderaciones['quiz_pct'],
                    'informe_pct': ponderaciones['informe_pct']
                }
            )
            DATOS['bloqueA'][estudiante] = nuevo_bloqueA
            
            bloqueB = DATOS['bloqueB_raw'].get(estudiante, 0)
            DATOS['final'][estudiante] = calcular_nota_final(
                nuevo_bloqueA, bloqueB,
                ponderaciones['bloqueA_pct'],
                ponderaciones['bloqueB_pct']
            )
        
        guardar_datos_json()
        return jsonify({'success': True, 'mensaje': 'Notas recalculadas correctamente'})
    except Exception as e:
        return jsonify({'error': str(e)}), 400

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)
